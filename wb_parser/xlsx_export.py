import json
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill


COLUMNS = [
    ("url", "Ссылка на товар"),
    ("article", "Артикул"),
    ("name", "Название"),
    ("price", "Цена (руб.)"),
    ("description", "Описание"),
    ("images", "Ссылки на изображения"),
    ("characteristics", "Характеристики"),
    ("seller_name", "Название селлера"),
    ("seller_url", "Ссылка на селлера"),
    ("sizes", "Размеры"),
    ("stock", "Остатки"),
    ("rating", "Рейтинг"),
    ("reviews_count", "Количество отзывов"),
]

FILTER_MIN_RATING = 4.5
FILTER_MAX_PRICE = 10000
FILTER_COUNTRY = "Россия"


def filter_items(items: list) -> list:
    result = []
    for item in items:
        try:
            rating = float(item.get("rating") or 0)
        except (ValueError, TypeError):
            rating = 0

        if rating < FILTER_MIN_RATING:
            continue

        try:
            price = float(item.get("price") or 0)
        except (ValueError, TypeError):
            price = 0

        if price > FILTER_MAX_PRICE:
            continue

        characteristics = item.get("characteristics") or {}
        country = find_country(characteristics)
        if FILTER_COUNTRY.lower() not in country.lower():
            continue

        result.append(item)

    return result


def find_country(characteristics) -> str:
    country_keys = [
        "Страна производства",
        "Страна производителя",
        "Страна",
        "Country",
    ]

    if isinstance(characteristics, dict) and (
        "groups" in characteristics or "options" in characteristics
    ):
        for group in characteristics.get("groups", []):
            for option in group.get("options", []):
                name = str(option.get("name", ""))
                value = option.get("value", "")
                for country_key in country_keys:
                    if country_key.lower() in name.lower():
                        return str(value)

        for option in characteristics.get("options", []):
            name = str(option.get("name", ""))
            value = option.get("value", "")
            for country_key in country_keys:
                if country_key.lower() in name.lower():
                    return str(value)

        return ""

    if isinstance(characteristics, dict):
        for key, value in characteristics.items():
            for country_key in country_keys:
                if country_key.lower() in str(key).lower():
                    return str(value)

    return ""


def write_xlsx(items: list, path: Path, sheet_name: str):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_name

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="4472C4")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for col_index, (_, header_title) in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_index, value=header_title)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align

    ws.row_dimensions[1].height = 30

    for row_index, item in enumerate(items, start=2):
        for col_index, (field_name, _) in enumerate(COLUMNS, start=1):
            value = item.get(field_name, "")

            if field_name == "characteristics" and isinstance(value, (dict, list)):
                value = json.dumps(value, ensure_ascii=False)

            cell = ws.cell(row=row_index, column=col_index, value=value)
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    column_widths = {
        1: 40,
        2: 12,
        3: 40,
        4: 12,
        5: 60,
        6: 60,
        7: 60,
        8: 30,
        9: 35,
        10: 25,
        11: 10,
        12: 10,
        13: 15,
    }
    for col, width in column_widths.items():
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = width

    ws.freeze_panes = "A2"
    wb.save(path)


def export_xlsx_files(items: list, output_dir: Path):
    output_dir.mkdir(exist_ok=True)

    catalog_path = output_dir / "catalog.xlsx"
    write_xlsx(items, catalog_path, "Каталог")

    filtered_items = filter_items(items)
    filtered_path = output_dir / "filtered.xlsx"
    write_xlsx(filtered_items, filtered_path, "Выборка")

    return catalog_path, filtered_path, len(filtered_items)
